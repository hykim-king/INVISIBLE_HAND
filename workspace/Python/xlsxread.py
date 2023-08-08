import os
import openpyxl

os.chdir("..")
xlsx_folder = os.path.join('doc', 'xlsx')
# 파일명
file_list = [f for f in os.listdir(xlsx_folder) if f.endswith('.xlsx')]

# 엑셀 파일 열기
for file_name in file_list:
    workbook = openpyxl.load_workbook(xlsx_folder+"\\"+file_name)
    # 첫 번째 시트 선택
    sheet = workbook.active
    # 처음 5개 행 출력
    print("행 출력")
    for row in sheet.iter_rows(1, sheet.max_row, values_only=True):
        print(row)
    # 엑셀 파일 닫기
    workbook.close()
