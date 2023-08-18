import os
import openpyxl
import warnings
import numpy as np
import cx_Oracle

warnings.simplefilter("ignore")
os.chdir("..")
xlsx_folder = os.path.join('doc', 'xlsx')
# 파일명
file_list_xlsx = [f for f in os.listdir(xlsx_folder) if f.endswith('.xlsx')]

dsn = cx_Oracle.makedsn('192.168.0.123', 1521, 'xe')
db = cx_Oracle.connect('INVISIBLEHAND', '4321', dsn)
cursor = db.cursor()

def get_xlsx_column(xlsxname):
    work_arr = []
    for file_name in file_list_xlsx:
        if file_name.find(xlsxname) != -1:
            count = 0
            #print(file_name)
            workbook = openpyxl.load_workbook(xlsx_folder + "\\" + file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(5, sheet.max_row, values_only=True):
                work_arr.append(np.array(row[2:7], dtype=np.float64))
                work_arr.append(np.array(row[7:12], dtype=np.float64))
                work_arr.append(np.array(row[12:17], dtype=np.float64))
                work_arr.append(np.array(row[17:22], dtype=np.float64))
                work_arr.append(np.array(row[22:27], dtype=np.float64))
    return np.array(work_arr)



subCategory= ['총자본순이익률','총자본영업이익률','경영자본영업이익률','자기자본영업이익률','매출액영업이익률']
jejo_upjong = ['식료품','음료','섬유제품 의복제외','의복 의복액세서리 및 모피제품','가죽 가방 및 신발',
'목재 및 나무제품','펄프 종이 및 종이제품','인쇄 및 기록매체','화학물질 및 화학제품','의료용 물질 및 의약품',
'고무제품 및 플라스틱제품','비금속 광물제품','1차 금속','금속가공제품','전자부품 컴퓨터 영상 음향 및 통신장비',
'의료 정밀 광학기기 및 시계','전기장비','기타 기계 및 장비','자동차 및 트레일러','기타 운송장비',
'가구','기타 제품'
]


yearDate = [2017,2018,2019,2020,2021]
publicCompany = get_xlsx_column("2017~2021상장기업의_부가가치_분석")      #(12,96)
print(publicCompany)
print(publicCompany[0][0],publicCompany[0][4])


#삽입
count =0
for index in range(len(jejo_upjong)):
    for column in range(len(yearDate)):
        cursor.execute(
            "INSERT INTO PUBLICCOMPANY VALUES(TO_DATE('{0}','YYYY'),'{1}','{2}',{3},{4},{5},{6},{7})".format(
                yearDate[column], "제조업", jejo_upjong[index],
                publicCompany[count][0], publicCompany[count][1], publicCompany[count][2],
                publicCompany[count][3], publicCompany[count][4],

            )
        )
        count = count + 1
'''
#cursor.execute("COMMIT")
'''

