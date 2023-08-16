import os
import openpyxl
import warnings
import numpy as np
import cx_Oracle
warnings.simplefilter("ignore")
os.chdir("..")
xlsx_folder = os.path.join('doc', 'xlsx')
# 파일명
file_list_xlsx = [f for f in os.listdir(xlsx_folder) if f.endswith('.xlsx')]

dsn = cx_Oracle.makedsn('192.168.0.123', 1521, 'xe')
db = cx_Oracle.connect('INVISIBLEHAND', '4321', dsn)
cursor = db.cursor()

def get_b_jejo_xlsx_column(xlsxname):
    work_arr = []
    for file_name in file_list_xlsx:
        if file_name.find(xlsxname) != -1:
            count = 0
            #print(file_name)
            workbook = openpyxl.load_workbook(xlsx_folder + "\\" + file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(237, sheet.max_row, values_only=True):
                count = count + 1
                if count % 6 == 0:
                    work_arr.append(np.array(row[3:], dtype=np.float64))
    return np.array(work_arr)

def get_jejo_xlsx_column(xlsxname):
    work_arr = []
    for file_name in file_list_xlsx:
        if file_name.find(xlsxname) != -1:
            count = 0
            #print(file_name)
            workbook = openpyxl.load_workbook(xlsx_folder + "\\" + file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(51, 183, values_only=True):
                count = count + 1
                if count % 6 == 0:
                    try:
                        work_arr.append(np.array(row[3:], dtype=np.float64))
                    except:
                        work_arr.append(72.1)
    return np.array(work_arr)


def get_jejo_x_xlsx_column(start_row,xlsxname):
    work_arr = []
    for file_name in file_list_xlsx:
        if file_name.find(xlsxname) != -1:
            # 엑셀 파일 불러오기
            workbook = openpyxl.load_workbook(xlsx_folder + "\\" + file_name)
            sheet = workbook.active

            # 행 단위로 데이터 읽기
            for row_index in range(start_row, start_row+22):  # 행의 범위 수정 가능
                row_data = []
                for col_index in range(9, sheet.max_column + 1, 6):
                    cell_value = sheet.cell(row=row_index, column=col_index).value
                    row_data.append(cell_value)
                work_arr.append(row_data)
    return np.array(work_arr)


B_jejo_upjong = ['건설업','서비스업','도매 및 소매업','운수업','숙박 및 음식점업',
'출판,영상,방송통신 및 정보서비스업','부동산업 및 임대업','전문,과학 및 기술서비스업','사업시설관리 및 사업지원서비스업','교육서비스업',
'예술,스포츠 및 여가관련서비스업','수리 및 기타개인서비스업'
]

jejo_upjong = ['식료품','음료','섬유제품 의복제외','의복 의복액세서리 및 모피제품','가죽 가방 및 신발',
'목재 및 나무제품','펄프 종이 및 종이제품','인쇄 및 기록매체','화학물질 및 화학제품','의료용 물질 및 의약품',
'고무제품 및 플라스틱제품','비금속 광물제품','1차 금속','금속가공제품','전자부품 컴퓨터 영상 음향 및 통신장비',
'의료 정밀 광학기기 및 시계','전기장비','기타 기계 및 장비','자동차 및 트레일러','기타 운송장비',
'가구','기타 제품'
]
monthdate = [201501, 201502, 201503, 201504, 201505, 201506, 201507, 201508, 201509, 201510, 201511, 201512,
             201601, 201602, 201603, 201604, 201605, 201606, 201607, 201608, 201609, 201610, 201611, 201612,
             201701, 201702, 201703, 201704, 201705, 201706, 201707, 201708, 201709, 201710, 201711, 201712,
             201801, 201802, 201803, 201804, 201805, 201806, 201807, 201808, 201809, 201810, 201811, 201812,
             201901, 201902, 201903, 201904, 201905, 201906, 201907, 201908, 201909, 201910, 201911, 201912,
             202001, 202002, 202003, 202004, 202005, 202006, 202007, 202008, 202009, 202010, 202011, 202012,
             202101, 202102, 202103, 202104, 202105, 202106, 202107, 202108, 202109, 202110, 202111, 202112,
             202201, 202202, 202203, 202204, 202205, 202206, 202207, 202208, 202209, 202210, 202211, 202212]
monthdate2023 = [202301, 202302, 202303, 202304, 202305, 202306, 202307]




'''
#비제조업 삽입
for index in range(len(monthdate)):
    for column in range(len(upjong)):
        cursor.execute(
            "INSERT INTO SBHITABLE VALUES(TO_DATE('{0}','YYYYMM'),'{1}','{2}',{3},{4},{5},{6},{7},{8},{9},{10},{11},{12})".format(
                monthdate[index], "비제조업", B_jejo_upjong[column],
                geongijunban[column][index], goyongsujun[column][index], neasupanma[column][index],
                suchul[column][index], yeonupiic[column][index], jagumsajung[column][index],
                0, 0, 0, 0)
        )

cursor.execute("COMMIT")
print("비제조업SBHI값 삽입완료.")
'''


#####################여기는 제조업 삽입하는 곳 입니다.



geongijunban = get_jejo_x_xlsx_column(44,"2023경기전반_실적")      #(12,96)
goyongsujun = get_jejo_x_xlsx_column(43,"2023고용수준_실적")      #(12,96)
neasupanma = get_jejo_x_xlsx_column(43,"2023내수판매_실적")      #(12,96)
suchul = get_jejo_x_xlsx_column(43,"2023수출실적")      #(12,96)
yeonupiic = get_jejo_x_xlsx_column(43,"2023영업이익_실적")      #(12,96)
jagumsajung = get_jejo_x_xlsx_column(43,"2023자금사정_실적")      #(12,96)



#제조업 삽입
for index in range(len(monthdate2023)):
    for column in range(len(B_jejo_upjong)):
        print(
            "INSERT INTO SBHITABLE VALUES(TO_DATE('{0}','YYYYMM'),'{1}','{2}',{3},{4},{5},{6},{7},{8},{9},{10},{11},{12});".format(
                monthdate2023[index], "비제조업", B_jejo_upjong[column],
                geongijunban[column][index], goyongsujun[column][index], neasupanma[column][index],
                suchul[column][index], yeonupiic[column][index], jagumsajung[column][index],
                0,0,0,0)
        )
#cursor.execute("COMMIT")
print("제조업SBHI값 삽입완료.")
